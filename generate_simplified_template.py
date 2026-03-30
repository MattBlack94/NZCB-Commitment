"""
Generate simplified NZCB Commitment Reporting Form Excel templates.

Produces two files:
  1. NZCB_Simplified_Standard.xlsx  (Sections 1-5)
  2. NZCB_Simplified_EP100.xlsx     (Sections 1-6, includes EP100 tab)

Based on FORM-SIMPLIFICATION-PROPOSAL.md.
"""

import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
NORMAL_FONT = Font(size=11)
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1B5E20")
SUBTITLE_FONT = Font(bold=True, size=13, color="1B5E20")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Dropdown lists (will be written to the hidden Lists tab)
# ---------------------------------------------------------------------------
LISTS = {
    "CommitmentVersion": ["2019", "2021"],
    "YesNo": ["Yes", "No"],
    "TargetYear": [str(y) for y in range(2025, 2051)],
    "AreaMeasurement": ["GFA", "GLA", "NLA", "NIA", "CFA", "Other"],
    "Coverage": ["Common areas only", "Whole building"],
    "AssetType": [
        "Residential",
        "Tenanted space",
        "Whole building (single)",
        "Whole building (multi-building/campus)",
        "New development / major renovation",
    ],
    "PortfolioChangeReason": [
        "Acquisitions",
        "Disposals",
        "New development completed",
        "Reclassification",
        "Boundary change",
        "Other",
    ],
    "GHGStandard": ["GHG Protocol", "ISO 14064", "NABERS", "Other"],
    "NonMeteredSource": [
        "Engineering estimates",
        "Benchmarking / proxy data",
        "Energy modelling",
        "Utility bills (estimated)",
        "Other",
    ],
    "VerificationMethod": [
        "Third-party certification",
        "Third-party verification",
        "SME self-assessment",
        "Other",
    ],
    "CompletionYear": [str(y) for y in range(2015, 2051)],
    "BuildingType": [
        "Office",
        "Retail",
        "Residential (multi-unit)",
        "Residential (single)",
        "Industrial / Warehouse",
        "Education",
        "Healthcare",
        "Hospitality",
        "Mixed use",
        "Other",
    ],
    "CalculationTool": [
        "One Click LCA",
        "eTool",
        "Tally",
        "Athena Impact Estimator",
        "RICS Whole Life Carbon",
        "Custom / In-house",
        "Other",
    ],
    "ECDatabase": [
        "WLCN / Carbon Heroes",
        "CLF Embodied Carbon Benchmark Database",
        "GBCA",
        "National database",
        "Other",
        "Not submitted",
    ],
    "ActionCategory": [
        "Energy Efficiency",
        "Electrification",
        "Fuel Switching",
        "Renewable (on-site)",
        "Renewable (off-site)",
        "Embodied Carbon Reduction",
        "Carbon Offsets",
        "Other",
    ],
    "ActionStatus": ["Not started", "In progress", "Completed"],
    "ActionScope": ["Asset level", "Portfolio level"],
    "AdvocacyProgramme": ["Yes", "No", "In development"],
    "StakeholderType": [
        "Tenants",
        "Owners/Developers",
        "Manufacturers",
        "Policymakers",
        "Customers/Clients",
        "Contractors",
        "Other",
    ],
    "AdvocacyActionType": [
        "Tenant support",
        "Sustainability services",
        "Supplier engagement",
        "Membership/partnership",
        "Events/speaking",
        "Embodied carbon",
        "Whole life carbon",
        "Other",
    ],
    "EP100CommitmentType": [
        "Energy Productivity",
        "Net Zero Carbon Buildings",
        "Cooling",
        "Multiple",
    ],
}


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _set_col_widths(ws, widths: dict):
    """Set column widths.  widths = {col_letter_or_num: width}."""
    for col, w in widths.items():
        letter = get_column_letter(col) if isinstance(col, int) else col
        ws.column_dimensions[letter].width = w


def _write_row(ws, row, values, font=None, fill=None, alignment=None, border=None,
               merge_to=None):
    """Write a list of values to a row, applying optional styles."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
    if merge_to:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)


def _style_input_cell(ws, row, col, fill=INPUT_FILL, border=THIN_BORDER):
    """Apply input-cell styling."""
    cell = ws.cell(row=row, column=col)
    cell.fill = fill
    cell.border = border
    return cell


def _add_dv(ws, formula, cell_range, allow_blank=True):
    """Add a list data validation referencing the Lists sheet."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def _list_range(name):
    """Return a formula reference to a named range in the Lists sheet."""
    items = LISTS[name]
    return f"Lists!$A$1:$A${len(items)}"  # placeholder; we build per-column refs


def _list_formula(col_letter, count):
    """Return absolute range formula for a column in the Lists sheet."""
    return f"=Lists!${col_letter}$2:${col_letter}${count + 1}"


def _header_row(ws, row, values, widths=None):
    """Write a dark-green header row and optionally set widths."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER


def _section_label(ws, row, text, cols=6):
    """Write a section header row spanning multiple columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = SECTION_FILL


def _label_input_pair(ws, row, label, input_col=2, label_col=1, span_input=1):
    """Write a label in col A and style input cell(s) starting at input_col."""
    cell = ws.cell(row=row, column=label_col, value=label)
    cell.font = BOLD_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for c in range(input_col, input_col + span_input):
        _style_input_cell(ws, row, c)


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def build_lists_tab(wb):
    """Create the hidden Lists tab with all dropdown source data."""
    ws = wb.create_sheet("Lists")

    # Write each list as a column with header in row 1
    col = 1
    list_columns = {}  # name -> (col_letter, item_count)
    for name, items in LISTS.items():
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=name).font = BOLD_FONT
        for i, item in enumerate(items, start=2):
            ws.cell(row=i, column=col, value=item)
        list_columns[name] = (letter, len(items))
        ws.column_dimensions[letter].width = max(len(name), max((len(str(x)) for x in items), default=10)) + 2
        col += 1

    ws.sheet_state = "hidden"
    return list_columns  # needed for data-validation formulas


def build_instructions_tab(wb):
    """Create the Instructions tab."""
    ws = wb.create_sheet("Instructions", 0)

    row = 2
    ws.merge_cells("A2:F2")
    cell = ws.cell(row=row, column=1,
                   value="WorldGBC Net Zero Carbon Buildings Commitment - Reporting Form 2025")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")

    row = 4
    ws.merge_cells("A4:F4")
    ws.cell(row=row, column=1, value="Instructions").font = SUBTITLE_FONT

    instructions = [
        "",
        "This workbook contains the simplified reporting form for the WorldGBC Net Zero Carbon Buildings Commitment.",
        "Please complete all applicable sections. The form is organised into the following sections:",
        "",
        "  Section 1 - Your Details: Confirm your entity information, contact details, and consent.",
        "  Section 2 - Your Portfolio: Describe your portfolio of assets, areas, and any changes from last year.",
        "  Section 3 - Performance Data: Report operational carbon (3A) and, if applicable, embodied carbon (3B).",
        "  Section 4 - Decarbonisation Actions: Describe your decarbonisation strategy and specific actions.",
        "  Section 5 - Advocacy: Describe your advocacy activities with stakeholders.",
        "  Section 6 - EP100: (EP100 members only) Report EP100-specific data.",
        "",
        "General guidance:",
        "  - Yellow cells are input fields. Please enter your data in these cells only.",
        "  - Dropdown fields have pre-defined options; click the cell to see the list.",
        "  - Where a section is marked as conditional, complete it only if it applies to you.",
        "  - All areas should be reported in square metres (m2).",
        "  - All energy should be reported in MWh.",
        "  - All emissions should be reported in tCO2e.",
        "",
        "For support, please contact:",
        "  WorldGBC Advancing Net Zero team",
        "  Email: nzcb@worldgbc.org",
        "  Website: https://worldgbc.org/advancing-net-zero",
    ]

    for i, line in enumerate(instructions):
        r = row + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=line)
        c.font = NORMAL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")

    _set_col_widths(ws, {1: 4, 2: 20, 3: 20, 4: 20, 5: 20, 6: 20})
    # Actually set col A wider to hold the text since we merge A:F
    ws.column_dimensions["A"].width = 100

    ws.sheet_properties.tabColor = "1B5E20"


def build_details_tab(wb, lc):
    """Section 1: Your Details."""
    ws = wb.create_sheet("1. Details")
    ws.sheet_properties.tabColor = "1B5E20"
    _set_col_widths(ws, {1: 35, 2: 40, 3: 20})

    row = 1
    _section_label(ws, row, "Section 1: Your Details", cols=3)

    fields = [
        ("Entity name", None),
        ("Signatory ID", None),
        ("Year joined", None),
        ("Commitment version", "CommitmentVersion"),
        ("Submitter name", None),
        ("Submitter email", None),
    ]
    row = 3
    for label, list_name in fields:
        _label_input_pair(ws, row, label, input_col=2)
        if label == "Entity name":
            ws.cell(row=row, column=2).value = "[Pre-fill entity name]"
        if list_name:
            formula = f"=Lists!${lc[list_name][0]}$2:${lc[list_name][0]}${lc[list_name][1] + 1}"
            _add_dv(ws, formula, f"B{row}")
        row += 1

    # Climate initiatives
    row += 1
    _section_label(ws, row, "Climate Initiative Memberships", cols=3)
    row += 1
    initiatives = ["EP100", "SBTi", "Race to Zero", "The Climate Pledge"]
    yn_formula = f"=Lists!${lc['YesNo'][0]}$2:${lc['YesNo'][0]}${lc['YesNo'][1] + 1}"
    for init_name in initiatives:
        _label_input_pair(ws, row, init_name, input_col=2)
        _add_dv(ws, yn_formula, f"B{row}")
        row += 1

    # Consent
    row += 1
    _section_label(ws, row, "Consent and Confirmation", cols=3)
    row += 1

    consent_text = (
        "I consent to WorldGBC and the Climate Group collecting, storing, and "
        "publishing the data provided in this form for the purposes of tracking "
        "progress under the Net Zero Carbon Buildings Commitment."
    )
    _label_input_pair(ws, row, consent_text, input_col=2)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 45
    _add_dv(ws, yn_formula, f"B{row}")
    row += 1

    _label_input_pair(ws, row, "Marketing opt-in (may WorldGBC use your data in public communications?)", input_col=2)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30
    _add_dv(ws, yn_formula, f"B{row}")
    row += 1

    _label_input_pair(ws, row, "Digital signature (type your full name)", input_col=2)

    ws.freeze_panes = "A2"


def build_portfolio_tab(wb, lc):
    """Section 2: Your Portfolio."""
    ws = wb.create_sheet("2. Portfolio")
    ws.sheet_properties.tabColor = "1B5E20"
    _set_col_widths(ws, {1: 45, 2: 20, 3: 20, 4: 25, 5: 30})

    row = 1
    _section_label(ws, row, "Section 2: Your Portfolio", cols=5)

    # Portfolio-level fields
    row = 3
    _label_input_pair(ws, row, "Target year", input_col=2)
    ty_formula = f"=Lists!${lc['TargetYear'][0]}$2:${lc['TargetYear'][0]}${lc['TargetYear'][1] + 1}"
    _add_dv(ws, ty_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "Area measurement type", input_col=2)
    am_formula = f"=Lists!${lc['AreaMeasurement'][0]}$2:${lc['AreaMeasurement'][0]}${lc['AreaMeasurement'][1] + 1}"
    _add_dv(ws, am_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "Coverage", input_col=2)
    cov_formula = f"=Lists!${lc['Coverage'][0]}$2:${lc['Coverage'][0]}${lc['Coverage'][1] + 1}"
    _add_dv(ws, cov_formula, f"B{row}")

    # Asset table
    row += 2
    _section_label(ws, row, "Asset Portfolio", cols=5)
    row += 1
    headers = ["Asset Type", "Number of assets", "Total area (m2)",
               "Change from last year (+/-)", "Reason for change"]
    _header_row(ws, row, headers)
    header_row_num = row

    at_formula = f"=Lists!${lc['AssetType'][0]}$2:${lc['AssetType'][0]}${lc['AssetType'][1] + 1}"
    pcr_formula = f"=Lists!${lc['PortfolioChangeReason'][0]}$2:${lc['PortfolioChangeReason'][0]}${lc['PortfolioChangeReason'][1] + 1}"

    asset_types = LISTS["AssetType"]
    for i, at in enumerate(asset_types):
        row += 1
        ws.cell(row=row, column=1, value=at).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for c in range(2, 6):
            _style_input_cell(ws, row, c)
        _add_dv(ws, pcr_formula, f"E{row}")

    # Portfolio changes section
    row += 2
    _section_label(ws, row, "Portfolio Changes", cols=5)
    row += 1
    changes = ["Assets added (count)", "Assets removed (count)", "Reason for changes"]
    for label in changes:
        _label_input_pair(ws, row, label, input_col=2)
        if label == "Reason for changes":
            _add_dv(ws, pcr_formula, f"B{row}")
        row += 1
    _label_input_pair(ws, row, "Additional details on changes", input_col=2)

    ws.freeze_panes = f"A{header_row_num + 1}"


def build_performance_tab(wb, lc):
    """Section 3: Performance Data (3A Operational + 3B Embodied)."""
    ws = wb.create_sheet("3. Performance")
    ws.sheet_properties.tabColor = "1B5E20"
    _set_col_widths(ws, {1: 35, 2: 12, 3: 22, 4: 22, 5: 22, 6: 18, 7: 20, 8: 20})

    row = 1
    _section_label(ws, row, "Section 3A: Operational Carbon", cols=6)

    # Reporting period
    row = 3
    _label_input_pair(ws, row, "Reporting period - From (DD/MM/YYYY)", input_col=2)
    row += 1
    _label_input_pair(ws, row, "Reporting period - To (DD/MM/YYYY)", input_col=2)
    row += 1
    _label_input_pair(ws, row, "GHG accounting standard", input_col=2)
    ghg_formula = f"=Lists!${lc['GHGStandard'][0]}$2:${lc['GHGStandard'][0]}${lc['GHGStandard'][1] + 1}"
    _add_dv(ws, ghg_formula, f"B{row}")

    # Data table
    row += 2
    _section_label(ws, row, "Operational Carbon Data", cols=6)
    row += 1
    headers = ["Metric", "Unit", "This Reporting Period",
               "Renewable (on-site)", "Renewable (off-site)", "Offsets"]
    _header_row(ws, row, headers)
    freeze_row = row + 1

    metrics = [
        ("Gross floor area", "m2", [True, False, False, False]),
        ("Electricity", "MWh", [True, True, True, False]),
        ("Fuels", "MWh", [True, False, False, False]),
        ("District heating/cooling", "MWh", [True, False, False, False]),
        ("Scope 1", "tCO2e", [True, False, False, True]),
        ("Scope 2 (location-based)", "tCO2e", [True, False, False, False]),
        ("Scope 2 (market-based)", "tCO2e", [True, False, False, False]),
    ]
    for metric_name, unit, input_cols in metrics:
        row += 1
        ws.cell(row=row, column=1, value=metric_name).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        for ci, active in enumerate(input_cols):
            col_num = ci + 3
            if active:
                _style_input_cell(ws, row, col_num)
            else:
                ws.cell(row=row, column=col_num).border = THIN_BORDER

    # Refrigerant emissions
    row += 2
    yn_formula = f"=Lists!${lc['YesNo'][0]}$2:${lc['YesNo'][0]}${lc['YesNo'][1] + 1}"
    _label_input_pair(ws, row, "Refrigerant emissions included?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")
    row += 1
    _label_input_pair(ws, row, "Refrigerant emissions quantity (tCO2e)", input_col=2)

    # Metered data
    row += 2
    _label_input_pair(ws, row, "% of data from metered sources", input_col=2)
    row += 1
    _label_input_pair(ws, row, "Primary non-metered data source", input_col=2)
    nm_formula = f"=Lists!${lc['NonMeteredSource'][0]}$2:${lc['NonMeteredSource'][0]}${lc['NonMeteredSource'][1] + 1}"
    _add_dv(ws, nm_formula, f"B{row}")

    # Verification
    row += 2
    _section_label(ws, row, "Verification", cols=6)
    row += 1
    _label_input_pair(ws, row, "Verification method", input_col=2)
    vm_formula = f"=Lists!${lc['VerificationMethod'][0]}$2:${lc['VerificationMethod'][0]}${lc['VerificationMethod'][1] + 1}"
    _add_dv(ws, vm_formula, f"B{row}")
    row += 1
    _label_input_pair(ws, row, "Verifier name", input_col=2)
    row += 1
    _label_input_pair(ws, row, "Verification date", input_col=2)
    row += 1
    _label_input_pair(ws, row, "Link to verification report", input_col=2)
    row += 1
    _label_input_pair(ws, row, "Public disclosure link (URL)", input_col=2)

    # -------------------------------------------------------------------
    # Section 3B: Embodied Carbon
    # -------------------------------------------------------------------
    row += 3
    _section_label(ws, row, "Section 3B: Embodied Carbon", cols=8)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    note = ws.cell(row=row, column=1,
                   value="Complete this section only if you have updated to 2021 commitment AND have new developments")
    note.font = Font(bold=True, italic=True, size=11, color="B71C1C")
    note.alignment = Alignment(wrap_text=True)

    row += 2
    ec_headers = ["Asset name", "Completion year", "Building type", "Area (m2)",
                  "Upfront carbon A1-A5 (tCO2e)", "Calculation tool",
                  "Database submitted to", "LCA conducted?"]
    _header_row(ws, row, ec_headers)

    cy_formula = f"=Lists!${lc['CompletionYear'][0]}$2:${lc['CompletionYear'][0]}${lc['CompletionYear'][1] + 1}"
    bt_formula = f"=Lists!${lc['BuildingType'][0]}$2:${lc['BuildingType'][0]}${lc['BuildingType'][1] + 1}"
    ct_formula = f"=Lists!${lc['CalculationTool'][0]}$2:${lc['CalculationTool'][0]}${lc['CalculationTool'][1] + 1}"
    db_formula = f"=Lists!${lc['ECDatabase'][0]}$2:${lc['ECDatabase'][0]}${lc['ECDatabase'][1] + 1}"

    for i in range(10):
        row += 1
        for c in range(1, 9):
            _style_input_cell(ws, row, c)
        _add_dv(ws, cy_formula, f"B{row}")
        _add_dv(ws, bt_formula, f"C{row}")
        _add_dv(ws, ct_formula, f"F{row}")
        _add_dv(ws, db_formula, f"G{row}")
        _add_dv(ws, yn_formula, f"H{row}")

    ws.freeze_panes = f"A{freeze_row}"


def build_actions_tab(wb, lc):
    """Section 4: Decarbonisation Actions."""
    ws = wb.create_sheet("4. Actions")
    ws.sheet_properties.tabColor = "1B5E20"
    _set_col_widths(ws, {1: 25, 2: 30, 3: 30, 4: 30, 5: 15, 6: 16, 7: 22, 8: 18, 9: 22})

    row = 1
    _section_label(ws, row, "Section 4: Decarbonisation Actions", cols=9)

    row = 3
    headers = ["Action category", "Target", "Key milestones", "Current progress",
               "Status", "Scope", "Expected completion", "Energy saving (MWh)",
               "Carbon reduction (tCO2e)"]
    _header_row(ws, row, headers)
    freeze_row = row + 1

    ac_formula = f"=Lists!${lc['ActionCategory'][0]}$2:${lc['ActionCategory'][0]}${lc['ActionCategory'][1] + 1}"
    as_formula = f"=Lists!${lc['ActionStatus'][0]}$2:${lc['ActionStatus'][0]}${lc['ActionStatus'][1] + 1}"
    sc_formula = f"=Lists!${lc['ActionScope'][0]}$2:${lc['ActionScope'][0]}${lc['ActionScope'][1] + 1}"

    for i in range(15):
        row += 1
        for c in range(1, 10):
            _style_input_cell(ws, row, c)
        _add_dv(ws, ac_formula, f"A{row}")
        _add_dv(ws, as_formula, f"E{row}")
        _add_dv(ws, sc_formula, f"F{row}")

    # Additional questions
    row += 2
    _section_label(ws, row, "Additional Information", cols=9)
    yn_formula = f"=Lists!${lc['YesNo'][0]}$2:${lc['YesNo'][0]}${lc['YesNo'][1] + 1}"

    row += 1
    _label_input_pair(ws, row, "QA undertaken on plan?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")
    row += 1
    _label_input_pair(ws, row, "QA details", input_col=2, span_input=3)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    row += 1
    _label_input_pair(ws, row, "Electrification plans?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")
    row += 1
    _label_input_pair(ws, row, "Electrification details", input_col=2, span_input=3)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    row += 1
    _label_input_pair(ws, row, "Carbon offset type and source", input_col=2, span_input=3)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    ws.freeze_panes = f"A{freeze_row}"


def build_advocacy_tab(wb, lc):
    """Section 5: Advocacy."""
    ws = wb.create_sheet("5. Advocacy")
    ws.sheet_properties.tabColor = "1B5E20"
    _set_col_widths(ws, {1: 28, 2: 28, 3: 50, 4: 35})

    row = 1
    _section_label(ws, row, "Section 5: Advocacy", cols=4)

    row = 3
    ap_formula = f"=Lists!${lc['AdvocacyProgramme'][0]}$2:${lc['AdvocacyProgramme'][0]}${lc['AdvocacyProgramme'][1] + 1}"
    _label_input_pair(ws, row, "Advocacy programme in place?", input_col=2)
    _add_dv(ws, ap_formula, f"B{row}")

    row += 2
    _section_label(ws, row, "Advocacy Actions (up to 8)", cols=4)
    row += 1
    headers = ["Stakeholder type", "Action type", "Description", "Link (optional)"]
    _header_row(ws, row, headers)
    freeze_row = row + 1

    sh_formula = f"=Lists!${lc['StakeholderType'][0]}$2:${lc['StakeholderType'][0]}${lc['StakeholderType'][1] + 1}"
    at_formula = f"=Lists!${lc['AdvocacyActionType'][0]}$2:${lc['AdvocacyActionType'][0]}${lc['AdvocacyActionType'][1] + 1}"

    for i in range(8):
        row += 1
        for c in range(1, 5):
            _style_input_cell(ws, row, c)
        _add_dv(ws, sh_formula, f"A{row}")
        _add_dv(ws, at_formula, f"B{row}")

    row += 2
    _section_label(ws, row, "Additional Information", cols=4)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Any additional information about your commitment:")
    cell.font = BOLD_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=4)
    for r in range(row, row + 4):
        for c in range(1, 5):
            _style_input_cell(ws, r, c)

    ws.freeze_panes = f"A{freeze_row}"


def build_ep100_tab(wb, lc):
    """Section 6: EP100 (EP100 version only)."""
    ws = wb.create_sheet("6. EP100")
    ws.sheet_properties.tabColor = "1B5E20"
    _set_col_widths(ws, {1: 45, 2: 35, 3: 25})

    row = 1
    _section_label(ws, row, "Section 6: EP100", cols=3)

    yn_formula = f"=Lists!${lc['YesNo'][0]}$2:${lc['YesNo'][0]}${lc['YesNo'][1] + 1}"
    ep_formula = f"=Lists!${lc['EP100CommitmentType'][0]}$2:${lc['EP100CommitmentType'][0]}${lc['EP100CommitmentType'][1] + 1}"

    row = 3
    _section_label(ws, row, "EP100 Reporting", cols=3)

    row += 1
    _label_input_pair(ws, row, "Reporting at group level?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "EP100 commitment type", input_col=2)
    _add_dv(ws, ep_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "EP100 target year", input_col=2)
    ty_formula = f"=Lists!${lc['TargetYear'][0]}$2:${lc['TargetYear'][0]}${lc['TargetYear'][1] + 1}"
    _add_dv(ws, ty_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "EP100 baseline year", input_col=2)

    row += 2
    _section_label(ws, row, "Financial Savings from Energy Efficiency", cols=3)

    row += 1
    _label_input_pair(ws, row, "Currency", input_col=2)

    row += 1
    _label_input_pair(ws, row, "Annual financial savings", input_col=2)

    row += 1
    _label_input_pair(ws, row, "Cumulative financial savings since baseline", input_col=2)

    row += 2
    _section_label(ws, row, "Heating and Cooling", cols=3)

    row += 1
    _label_input_pair(ws, row, "% of heating from electricity", input_col=2)

    row += 1
    _label_input_pair(ws, row, "% of cooling from electricity", input_col=2)

    row += 1
    _label_input_pair(ws, row, "% of heating from renewable sources", input_col=2)

    row += 1
    _label_input_pair(ws, row, "% of cooling from renewable sources", input_col=2)

    row += 1
    _label_input_pair(ws, row, "District heating/cooling used?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")

    row += 2
    _section_label(ws, row, "Additional EP100 Information", cols=3)
    row += 1
    _label_input_pair(ws, row, "Energy management system in place?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "ISO 50001 certified?", input_col=2)
    _add_dv(ws, yn_formula, f"B{row}")

    row += 1
    _label_input_pair(ws, row, "Additional notes", input_col=2)

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main workbook assembly
# ---------------------------------------------------------------------------

def create_workbook(include_ep100: bool) -> Workbook:
    """Build a complete workbook, optionally including the EP100 tab."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build Lists tab first (needed for data validation references)
    lc = build_lists_tab(wb)

    # Build content tabs
    build_instructions_tab(wb)
    build_details_tab(wb, lc)
    build_portfolio_tab(wb, lc)
    build_performance_tab(wb, lc)
    build_actions_tab(wb, lc)
    build_advocacy_tab(wb, lc)

    if include_ep100:
        build_ep100_tab(wb, lc)

    return wb


def main():
    import os

    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Standard version (Sections 1-5)
    wb_standard = create_workbook(include_ep100=False)
    standard_path = os.path.join(base_dir, "NZCB_Simplified_Standard.xlsx")
    wb_standard.save(standard_path)
    print(f"Created: {standard_path}")

    # EP100 version (Sections 1-6)
    wb_ep100 = create_workbook(include_ep100=True)
    ep100_path = os.path.join(base_dir, "NZCB_Simplified_EP100.xlsx")
    wb_ep100.save(ep100_path)
    print(f"Created: {ep100_path}")


if __name__ == "__main__":
    main()
